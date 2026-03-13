import streamlit as st
import io
import re
import calendar
from collections import OrderedDict
from datetime import date

import pdfplumber
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Xcel Bill Processor",
    page_icon="⚡",
    layout="centered",
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono&display=swap');

    html, body, [class*="css"] {
        font-family: 'DM Sans', sans-serif;
    }
    .main { background: #f7f8fa; }
    .block-container { max-width: 720px; padding-top: 2rem; }

    .hero {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 60%, #0f3460 100%);
        border-radius: 16px;
        padding: 2.5rem 2rem;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 8px 32px rgba(0,0,0,0.18);
    }
    .hero h1 {
        color: #ffffff;
        font-size: 2rem;
        font-weight: 600;
        margin: 0 0 0.5rem 0;
        letter-spacing: -0.5px;
    }
    .hero p {
        color: #a0aec0;
        font-size: 1rem;
        margin: 0;
    }
    .hero .bolt { font-size: 2.5rem; margin-bottom: 0.75rem; display: block; }

    .step-card {
        background: white;
        border-radius: 12px;
        padding: 1.5rem;
        margin-bottom: 1rem;
        border: 1px solid #e2e8f0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.04);
    }
    .step-num {
        display: inline-block;
        background: #0f3460;
        color: white;
        border-radius: 50%;
        width: 28px; height: 28px;
        text-align: center;
        line-height: 28px;
        font-size: 0.85rem;
        font-weight: 600;
        margin-right: 0.5rem;
    }

    .stDownloadButton > button {
        background: linear-gradient(135deg, #0f3460, #e94560) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 0.75rem 2rem !important;
        font-size: 1rem !important;
        font-weight: 600 !important;
        width: 100% !important;
        margin-top: 0.5rem !important;
        box-shadow: 0 4px 15px rgba(233,69,96,0.3) !important;
        transition: all 0.2s !important;
    }
    .stDownloadButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(233,69,96,0.4) !important;
    }

    .success-box {
        background: #f0fff4;
        border: 1px solid #68d391;
        border-radius: 10px;
        padding: 1rem 1.25rem;
        margin: 1rem 0;
        color: #276749;
        font-weight: 500;
    }
    .warn-box {
        background: #fffbeb;
        border: 1px solid #f6ad55;
        border-radius: 10px;
        padding: 1rem 1.25rem;
        margin: 1rem 0;
        color: #744210;
    }
    .stat-row {
        display: flex;
        gap: 1rem;
        margin: 1rem 0;
    }
    .stat {
        flex: 1;
        background: white;
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
        border: 1px solid #e2e8f0;
    }
    .stat-val { font-size: 1.6rem; font-weight: 700; color: #0f3460; }
    .stat-lbl { font-size: 0.78rem; color: #718096; margin-top: 2px; }
    .footer {
        text-align: center;
        color: #a0aec0;
        font-size: 0.8rem;
        margin-top: 2rem;
        padding-top: 1rem;
        border-top: 1px solid #e2e8f0;
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# 0. Helpers
# ─────────────────────────────────────────────

def parse_money(s):
    if s is None:
        return None
    cleaned = str(s).replace("$", "").replace(",", "").replace("CR", "").replace("-", "").strip()
    return float(cleaned) if cleaned else None

def parse_date(mmddyy):
    m, d, y = mmddyy.split("/")
    return date(2000 + int(y), int(m), int(d))

def month_key(mmddyy):
    m, _, y = mmddyy.split("/")
    return f"{int(m)}/{y}"

def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def safe_save(wb, path):
    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = path.with_name(f"{path.stem}_{int(time.time())}{path.suffix}")
        wb.save(alt)
        return alt

def prorate_amount(start_str, end_str, amount):
    """
    Split amount across calendar months proportionally by days.
    Special rule: Dec->Jan always goes fully into January.
    Returns OrderedDict {month_key: amount}.
    """
    s = parse_date(start_str)
    e = parse_date(end_str)
    total_days = (e - s).days
    start_mk = month_key(start_str)
    end_mk   = month_key(end_str)

    if total_days <= 0 or start_mk == end_mk:
        return OrderedDict([(start_mk, round(amount, 2))])

    # December -> January: full amount into January
    if s.month == 12 and e.month == 1:
        return OrderedDict([(end_mk, round(amount, 2))])

    last_day_of_start = date(s.year, s.month, calendar.monthrange(s.year, s.month)[1])
    days_in_start = (min(last_day_of_start, e) - s).days
    amt_start = round(amount * days_in_start / total_days, 2)
    amt_end   = round(amount - amt_start, 2)

    result = OrderedDict()
    if amt_start:
        result[start_mk] = amt_start
    if amt_end:
        result[end_mk] = round(result.get(end_mk, 0.0) + amt_end, 2)
    return result

def add_proration(allocations, start_str, end_str, amount):
    for mk, val in prorate_amount(start_str, end_str, amount).items():
        allocations[mk] = round(allocations.get(mk, 0.0) + val, 2)


# ─────────────────────────────────────────────
# 1. Summary parsing
# ─────────────────────────────────────────────

# Captures: (premise#) (identifier) (optional -) $(amount) (optional CR) (space or end)
# Barcode/QR junk after the amount is ignored.
_BILLED_RE     = re.compile(r"^(\d{9})\s+(.+?)\s+(-?)\$([\d,]+\.\d{2})\s*(CR)?(\s|$)", re.I)
_NOT_INC_RE    = re.compile(r"^(\d{9})\s+(.+?)\s+NOT\s*INCLUDED(\s|$)", re.I)
_DATE_IDENT_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
_PREM_NUM_RE   = re.compile(r"^(\d{9})\s+(.+)$")
_MONEY_RE      = re.compile(r"^(-?)\$([\d,]+\.\d{2})\s*(CR)?\s*$", re.I)
_NI_ONLY_RE    = re.compile(r"^NOT\s*INCLUDED\s*$", re.I)

def _clean_identifier(raw):
    return re.sub(r"\s+", " ", raw).strip().rstrip("_").strip()

def _bill_amount_from_match(m):
    """Extract signed dollar amount from a _BILLED_RE match."""
    neg = m.group(3) == "-"
    cr  = bool(m.group(5))
    amt = parse_money(m.group(4))
    if amt is not None and (neg or cr):
        amt = -amt
    return amt

def _bill_amount_from_next_line(nxt):
    """Extract signed dollar amount from a standalone money line, or None."""
    mm = _MONEY_RE.match(nxt)
    if not mm:
        return None, False
    neg = mm.group(1) == "-"
    cr  = bool(mm.group(3))
    amt = parse_money(mm.group(2))
    if amt is not None and (neg or cr):
        amt = -amt
    return amt, True

def parse_summary(pdf):
    summary = OrderedDict()
    for page in pdf.pages:
        text = page.extract_text() or ""
        if not re.search(r"PREMISES\s*SUMMARY", text, re.I):
            continue
        lines = [l.strip() for l in text.splitlines()]
        i = 0
        while i < len(lines):
            line = lines[i]

            # Fully-contained billed line (positive or credit)
            m = _BILLED_RE.match(line)
            if m:
                ident = m.group(2).strip()
                if not _DATE_IDENT_RE.match(ident):
                    summary[m.group(1)] = {
                        "premises_number":     m.group(1),
                        "premises_identifier": _clean_identifier(ident),
                        "current_bill":        _bill_amount_from_match(m),
                    }
                i += 1
                continue

            # Fully-contained NOT INCLUDED line
            m = _NOT_INC_RE.match(line)
            if m:
                summary[m.group(1)] = {
                    "premises_number":     m.group(1),
                    "premises_identifier": _clean_identifier(m.group(2)),
                    "current_bill":        None,
                }
                i += 1
                continue

            # Premise# + identifier with amount or NI on NEXT line
            m = _PREM_NUM_RE.match(line)
            if m:
                pnum  = m.group(1)
                ident = m.group(2).strip()
                if not _DATE_IDENT_RE.match(ident) and i + 1 < len(lines):
                    nxt = lines[i + 1].strip()
                    amt, matched = _bill_amount_from_next_line(nxt)
                    if matched:
                        summary[pnum] = {
                            "premises_number":     pnum,
                            "premises_identifier": _clean_identifier(ident),
                            "current_bill":        amt,
                        }
                        i += 2
                        continue
                    if _NI_ONLY_RE.match(nxt):
                        summary[pnum] = {
                            "premises_number":     pnum,
                            "premises_identifier": _clean_identifier(ident),
                            "current_bill":        None,
                        }
                        i += 2
                        continue
            i += 1
    return summary


# ─────────────────────────────────────────────
# 2. Detail section extraction
# ─────────────────────────────────────────────

_PREMISE_HDR_RE = re.compile(r"PREMISES?\s*NUMBER:\s*(\d{9})", re.I)

def extract_detail_sections(pdf):
    sections = OrderedDict()
    current  = None
    for page in pdf.pages:
        text = page.extract_text() or ""
        m = _PREMISE_HDR_RE.search(text)
        if m:
            current = m.group(1)
            sections[current] = text
        elif current:
            sections[current] += "\n" + text
    return sections


# ─────────────────────────────────────────────
# 3. Block extraction
# ─────────────────────────────────────────────

_DATE_RE = re.compile(
    r"Read\s*Dates:\s*(\d{2}/\d{2}/\d{2})\s*-\s*(\d{2}/\d{2}/\d{2})", re.I)
_SUB_RE  = re.compile(r"Sub\s*total\s*\$([\d,]+\.\d{2})", re.I)

def extract_blocks(text):
    flat = re.sub(r"\s+", " ", text)
    date_matches = list(_DATE_RE.finditer(flat))
    sub_matches  = list(_SUB_RE.finditer(flat))
    blocks  = []
    sub_idx = 0
    for i, dm in enumerate(date_matches):
        block_start = dm.end()
        block_end   = date_matches[i+1].start() if i+1 < len(date_matches) else len(flat)
        subtotal    = None
        while sub_idx < len(sub_matches):
            sm = sub_matches[sub_idx]
            if sm.start() < block_start:
                sub_idx += 1
            elif sm.start() < block_end:
                subtotal = parse_money(sm.group(1))
                sub_idx += 1
                break
            else:
                break
        blocks.append({
            "start_date": dm.group(1),
            "end_date":   dm.group(2),
            "subtotal":   subtotal,
        })
    return blocks

def extract_premises_total(text):
    flat = re.sub(r"\s+", " ", text)
    m = re.search(r"Premises\s*Total\s*\$([\d,]+\.\d{2})", flat, re.I)
    return parse_money(m.group(1)) if m else None


# ─────────────────────────────────────────────
# 4. Premise allocation
# ─────────────────────────────────────────────

def allocate(text):
    blocks         = extract_blocks(text)
    premises_total = extract_premises_total(text)
    allocations    = OrderedDict()

    if not blocks or premises_total is None:
        return allocations, premises_total, blocks

    # Single block: pro-rate across month boundary if needed
    if len(blocks) == 1:
        b = blocks[0]
        add_proration(allocations, b["start_date"], b["end_date"], premises_total)
        return allocations, premises_total, blocks

    # Multiple blocks (meter change)
    total_subtotals = sum(b["subtotal"] for b in blocks if b["subtotal"] is not None)

    if total_subtotals > premises_total + 0.005:
        # OVERFLOW: assign each block's amount to its start month (no pro-ration),
        # capped by remaining budget.
        budget = premises_total
        for block in blocks:
            if block["subtotal"] is None or budget < 0.005:
                break
            amount = min(block["subtotal"], round(budget, 2))
            mk = month_key(block["start_date"])
            allocations[mk] = round(allocations.get(mk, 0.0) + amount, 2)
            budget = round(budget - amount, 2)
        # Absorb rounding residual
        diff = round(premises_total - round(sum(allocations.values()), 2), 2)
        if abs(diff) >= 0.01 and allocations:
            last_key = list(allocations.keys())[-1]
            allocations[last_key] = round(allocations[last_key] + diff, 2)

    else:
        # NORMAL: each block's subtotal pro-rated across its month boundary;
        # remainder goes to end month of last block.
        subtotal_sum = 0.0
        for block in blocks:
            if block["subtotal"] is not None:
                add_proration(allocations, block["start_date"], block["end_date"], block["subtotal"])
                subtotal_sum += block["subtotal"]
        remainder = round(premises_total - subtotal_sum, 2)
        if abs(remainder) >= 0.01:
            end_mk = month_key(blocks[-1]["end_date"])
            allocations[end_mk] = round(allocations.get(end_mk, 0.0) + remainder, 2)

    return allocations, premises_total, blocks


# ─────────────────────────────────────────────
# 4b. Non-recurring charges parsing
# ─────────────────────────────────────────────

_NRC_SECTION_RE = re.compile(r"NON.RECURRING\s+CHARGES", re.I)
_NRC_ITEM_RE    = re.compile(r"^(.+?)\s+\$([\d,]+\.\d{2})(\s|$)")

def parse_non_recurring(pdf):
    """
    Find NON-RECURRING CHARGES/CREDITS SUMMARY and return list of
    {"description": str, "amount": float}. Skips header and total lines.
    Parses across ALL pages as one stream so page breaks don't reset state.
    """
    # Concatenate all page text first so page boundaries don't break parsing
    full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    lines = full_text.splitlines()

    charges = []
    in_section = False

    for line in lines:
        line = line.strip()
        if _NRC_SECTION_RE.search(line):
            in_section = True
            continue
        if not in_section:
            continue
        # Stop at next major section
        if re.match(
            r"(INFORMATION ABOUT YOUR BILL|PREMISES SUMMARY|ELECTRICITY SERVICE|^Page\s+\d)",
            line, re.I
        ):
            break
        # Skip header/empty/total lines
        if not line or re.match(r"(DESCRIPTION|CURRENT BILL)", line, re.I):
            continue
        if re.match(r"Total\s+\$", line, re.I):
            continue
        m = _NRC_ITEM_RE.match(line)
        if m:
            desc = m.group(1).strip()
            amt  = parse_money(m.group(2))
            if desc and amt and amt > 0:
                charges.append({"description": desc.title(), "amount": amt})

    # Deduplicate
    seen = set()
    unique = []
    for c in charges:
        key = (c["description"], c["amount"])
        if key not in seen:
            seen.add(key)
            unique.append(c)
    return unique

# ─────────────────────────────────────────────
# 5. Bill orchestration
# ─────────────────────────────────────────────

def parse_bill(pdf_bytes):
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        summary        = parse_summary(pdf)
        details        = extract_detail_sections(pdf)
        non_recurring  = parse_non_recurring(pdf)

    rows       = []
    all_months = set()

    for premise, info in summary.items():
        # Skip NOT INCLUDED (current_bill is None)
        if info["current_bill"] is None:
            continue

        row = {
            "premises_number":     premise,
            "premises_identifier": info["premises_identifier"],
            "current_bill":        info["current_bill"],
            "months":              OrderedDict(),
            "total":               None,
            "diff":                None,
            "notes":               "",
        }

        if info["current_bill"] < 0:
            row["notes"] = "CREDIT"
            if premise in details:
                _cr_blocks = extract_blocks(details[premise])
                if _cr_blocks:
                    mk = month_key(_cr_blocks[-1]["end_date"])
                    row["months"][mk] = info["current_bill"]
                    all_months.add(mk)
            row["total"] = info["current_bill"]
            row["diff"]  = 0.0
        elif premise in details:
            allocs, _, blocks = allocate(details[premise])
            row["months"] = allocs
            row["total"]  = round(sum(allocs.values()), 2) if allocs else None
            if row["total"] is not None:
                row["diff"] = round(info["current_bill"] - row["total"], 2)
            if len(blocks) > 1:
                row["notes"] = f"METER CHANGE {blocks[0]['end_date']}"
            elif details.get(premise, "") and "estimate" in details[premise].lower():
                row["notes"] = "ESTIMATED"
            all_months.update(allocs.keys())
        elif info["current_bill"] is not None:
            # Has a bill amount but no detail page — put full amount in notes
            row["notes"] = "NO DETAIL PAGE"

        rows.append(row)

    sorted_months = sorted(
        all_months,
        key=lambda x: (int(x.split("/")[1]), int(x.split("/")[0])),
    )
    return rows, sorted_months, non_recurring


def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

MONEY_FMT = '$#,##0.00_);($#,##0.00)'
HDR_FILL  = PatternFill("solid", start_color="1F1F1F")
HDR_FONT  = Font(color="FFFFFF", bold=True, name="Arial")
ROW_FILL  = PatternFill("solid", start_color="DCE6F1")
TOT_FILL  = PatternFill("solid", start_color="EDEDED")
NRC_FILL  = PatternFill("solid", start_color="FFF9E6")
YLW_FILL  = PatternFill("solid", start_color="FFF200")
BASE_FONT = Font(name="Arial")

def export_excel(rows, month_cols, statement_number, non_recurring=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xcel Allocations"

    headers = (
        ["PREMISES NUMBER", "PREMISES IDENTIFIER", "CURRENT BILL"]
        + month_cols
        + ["TOTAL", "DIFF", "NOTES"]
    )
    COL_CB    = 3
    COL_M0    = 4
    COL_TOTAL = COL_M0 + len(month_cols)
    COL_DIFF  = COL_TOTAL + 1
    COL_NOTES = COL_DIFF + 1

    # Header row
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

    # Data rows
    for ri, row in enumerate(rows, 2):
        ws.cell(ri, 1, row["premises_number"]).font = BASE_FONT
        ws.cell(ri, 2, row["premises_identifier"]).font = BASE_FONT

        c = ws.cell(ri, COL_CB, row["current_bill"])
        if row["current_bill"] is not None:
            c.number_format = MONEY_FMT

        for mi, mon in enumerate(month_cols):
            val = row["months"].get(mon)
            c = ws.cell(ri, COL_M0 + mi, val)
            if val is not None:
                c.number_format = MONEY_FMT

        c = ws.cell(ri, COL_TOTAL, row["total"])
        if row["total"] is not None:
            c.number_format = MONEY_FMT
            c.font = Font(bold=True, name="Arial")

        c = ws.cell(ri, COL_DIFF, row["diff"])
        if row["diff"] is not None:
            c.number_format = MONEY_FMT

        ws.cell(ri, COL_NOTES, row["notes"])

        for ci in range(1, COL_NOTES + 1):
            c        = ws.cell(ri, ci)
            c.border = thin_border()
            c.fill   = ROW_FILL

    # Non-recurring charge rows (late fees, etc.) — yellow highlight
    if non_recurring:
        for nrc in non_recurring:
            nrc_ri = ws.max_row + 1
            ws.cell(nrc_ri, 2, nrc["description"]).font = Font(bold=True, name="Arial")
            c = ws.cell(nrc_ri, COL_CB, nrc["amount"])
            c.number_format = MONEY_FMT
            c.font = Font(bold=True, name="Arial")
            # Place in last month column
            if month_cols:
                last_m_col = COL_M0 + len(month_cols) - 1
                c2 = ws.cell(nrc_ri, last_m_col, nrc["amount"])
                c2.number_format = MONEY_FMT
            c3 = ws.cell(nrc_ri, COL_TOTAL, nrc["amount"])
            c3.number_format = MONEY_FMT
            c3.font = Font(bold=True, name="Arial")
            ws.cell(nrc_ri, COL_NOTES, "NON-RECURRING CHARGE")
            for ci in range(1, COL_NOTES + 1):
                c        = ws.cell(nrc_ri, ci)
                c.border = thin_border()
                c.fill   = NRC_FILL

    # Totals row
    tr        = ws.max_row + 1
    last_data = tr - 1
    ws.cell(tr, 1, "Total")
    for ci in range(1, COL_NOTES + 1):
        c        = ws.cell(tr, ci)
        c.fill   = TOT_FILL
        c.border = thin_border()
        c.font   = Font(bold=True, name="Arial")
    for ci in range(COL_CB, COL_TOTAL + 1):
        col_ltr = get_column_letter(ci)
        c = ws.cell(tr, ci, f"=SUM({col_ltr}2:{col_ltr}{last_data})")
        c.number_format = MONEY_FMT
        c.font = Font(bold=True, name="Arial")

    # Statement number label
    nr = tr + 2
    ws.merge_cells(start_row=nr, start_column=2, end_row=nr, end_column=4)
    c           = ws.cell(nr, 2, statement_number)
    c.fill      = YLW_FILL
    c.font      = Font(bold=True, name="Arial")
    c.alignment = Alignment(horizontal="center")
    c.border    = thin_border()

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions[get_column_letter(COL_CB)].width = 14
    for mi in range(len(month_cols)):
        ws.column_dimensions[get_column_letter(COL_M0 + mi)].width = 12
    ws.column_dimensions[get_column_letter(COL_TOTAL)].width = 14
    ws.column_dimensions[get_column_letter(COL_DIFF)].width  = 12
    ws.column_dimensions[get_column_letter(COL_NOTES)].width = 24
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────

st.markdown("""
<div class="hero">
    <span class="bolt">⚡</span>
    <h1>Xcel Bill Processor</h1>
    <p>Upload an Xcel Energy PDF bill — get a monthly allocation Excel file instantly</p>
</div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader("Drop your Xcel Energy PDF here", type=["pdf"], label_visibility="collapsed")

if not uploaded:
    st.markdown("""
    <div class="step-card">
        <span class="step-num">1</span> <strong>Upload</strong> your Xcel Energy PDF bill above
    </div>
    <div class="step-card">
        <span class="step-num">2</span> <strong>Wait</strong> a few seconds while it processes
    </div>
    <div class="step-card">
        <span class="step-num">3</span> <strong>Download</strong> the Excel file with monthly allocations
    </div>
    """, unsafe_allow_html=True)
else:
    with st.spinner("Processing bill..."):
        try:
            pdf_bytes = uploaded.read()
            rows, month_cols, non_recurring = parse_bill(pdf_bytes)

            if not rows:
                st.markdown('<div class="warn-box">⚠️ No billed premises found in this PDF. Make sure it\'s an Xcel Energy bill.</div>', unsafe_allow_html=True)
            else:
                statement_number = uploaded.name.replace(".pdf", "").replace(".PDF", "")
                excel_bytes = export_excel(rows, month_cols, statement_number, non_recurring)

                total_bill = sum(r["current_bill"] for r in rows if r["current_bill"])
                meter_changes = sum(1 for r in rows if "METER CHANGE" in r.get("notes",""))

                st.markdown(f"""
                <div class="success-box">✅ Successfully processed <strong>{len(rows)}</strong> billed premises</div>
                <div class="stat-row">
                    <div class="stat">
                        <div class="stat-val">{len(rows)}</div>
                        <div class="stat-lbl">Billed Premises</div>
                    </div>
                    <div class="stat">
                        <div class="stat-val">{len(month_cols)}</div>
                        <div class="stat-lbl">Month Columns</div>
                    </div>
                    <div class="stat">
                        <div class="stat-val">${total_bill:,.2f}</div>
                        <div class="stat-lbl">Total Bill</div>
                    </div>
                    <div class="stat">
                        <div class="stat-val">{meter_changes}</div>
                        <div class="stat-lbl">Meter Changes</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.download_button(
                    label="⬇️  Download Excel File",
                    data=excel_bytes,
                    file_name=f"xcel_bill_{statement_number}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        except Exception as e:
            st.markdown(f'<div class="warn-box">❌ Error processing file: {str(e)}</div>', unsafe_allow_html=True)

st.markdown('<div class="footer">Xcel Bill Processor v7 · Forty Acres Energy</div>', unsafe_allow_html=True)
